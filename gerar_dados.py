
#!/usr/bin/env python3
"""
gerar_dados.py — Copa Bracell Absoluto
Converte controle_copa_bracell.xlsx → dados.json

Uso:
    python3 gerar_dados.py
    python3 gerar_dados.py --arquivo outro.xlsx --saida outro.json
    python3 gerar_dados.py --debug      (mostra cada célula lida)

Dependência:  pip install openpyxl
"""
import json
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  Instale a dependência:  pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────
#  MAPEAMENTO: coluna Excel → id da figurinha
#  DEVE ser idêntico ao array STICKERS no album_copa_bracell.html
# ─────────────────────────────────────────────────────────────────
STICKER_COLS = {
    5:  "efetividade_90",
    6:  "pe_50pct_prom",
    7:  "pe_100pct_prom",
    8:  "cross_50pct",
    9:  "terminal_20pct",
    10: "ilha_20pct",
    11: "display_turbilhao",
    12: "criativo_copa_2",
    13: "criativo_copa_4",
    14: "criativo_sj_2",
    15: "criativo_sj_4",
    16: "mpdv_leve_25",
    17: "mpdv_leve_50",
    18: "mpdv_pesado_1",
    19: "mpdv_pesado_2",
    20: "autofalante",
}

DATA_START_ROW = 4   # linha onde começam os dados das equipes
TEAM_ID_COL    = 4   # coluna D = ID da equipe


def cell_is_sim(value) -> bool:
    """True para SIM, sim, Sim, ' SIM ', etc."""
    if value is None:
        return False
    return str(value).strip().upper() == "SIM"

path = f'C:/Users/r.torres/OneDrive - SPOT/RAFAELA/Python/Campanha/controle_copa_bracell.xlsx'

def main():
    parser = argparse.ArgumentParser(
        description="Exporta controle_copa_bracell.xlsx para dados.json"
    )
    parser.add_argument("--arquivo", default=path)
    parser.add_argument("--saida",   default="dados.json")
    parser.add_argument("--debug",   action="store_true",
                        help="Imprime cada célula lida (útil para diagnóstico)")
    args = parser.parse_args()

    xlsx_path = Path(args.arquivo)
    if not xlsx_path.exists():
        print(f"❌  Arquivo não encontrado: {xlsx_path.resolve()}")
        print(f"    Certifique-se de rodar o script na mesma pasta da planilha.")
        sys.exit(1)

    print(f"📂  Lendo: {xlsx_path.resolve()}")

    # data_only=False garante leitura dos valores digitados diretamente
    # (data_only=True só lê fórmulas recalculadas, o que requer LibreOffice)
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        print(f"❌  Erro ao abrir a planilha: {e}")
        sys.exit(1)

    if "Controle" not in wb.sheetnames:
        print(f"❌  Aba 'Controle' não encontrada.")
        print(f"    Abas disponíveis: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["Controle"]
    resultado   = {"equipes": {}}
    total_sim   = 0
    total_cells = 0

    row = DATA_START_ROW
    while True:
        raw_id = ws.cell(row, TEAM_ID_COL).value
        if raw_id is None:
            break

        team_id   = str(raw_id).strip()
        team_data = {}

        if args.debug:
            print(f"\n  [{row}] Equipe: {repr(raw_id)} → '{team_id}'")

        for col, sid in STICKER_COLS.items():
            raw    = ws.cell(row, col).value
            earned = cell_is_sim(raw)
            team_data[sid] = earned
            total_cells   += 1
            if earned:
                total_sim += 1
            if args.debug:
                mark = "✅ SIM" if earned else "○    "
                print(f"    col {col:2d}  {sid:<22}  {repr(str(raw)):<14}  {mark}")

        resultado["equipes"][team_id] = team_data
        row += 1

    # ── Salvar JSON ──
    out_path = Path(args.saida)
    try:
        out_path.write_text(
            json.dumps(resultado, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception as e:
        print(f"❌  Erro ao salvar {out_path}: {e}")
        sys.exit(1)

    # ── Relatório ──
    n = len(resultado["equipes"])
    print(f"\n✅  {out_path.name} atualizado!")
    print(f"    Equipes: {n}  |  Figurinhas marcadas: {total_sim}/{total_cells}")
    print()
    print("─" * 54)
    for tid, st in resultado["equipes"].items():
        e   = sum(1 for v in st.values() if v)
        t   = len(st)
        bar = "█" * e + "░" * (t - e)
        pct = round(e / t * 100) if t else 0
        print(f"  {tid:<14}  [{bar}]  {e:>2}/{t}  ({pct}%)")
    print("─" * 54)
    print()
    print("💡  Recarregue o álbum ou clique em  ↻ Atualizar")


if __name__ == "__main__":
    main()